"""
readbom.py - read a BOM spreadsheet and pull out the part identifiers.

Accepts .xlsx / .xlsm / .csv. Finds the header row, then reports the columns
that look like manufacturer part numbers, distributor part numbers, LCSC codes
or product URLs, so a batch of parts can be worked through in one pass.

It does not guess a part's identity - it only surfaces what the sheet says.

Usage:
    python scripts/readbom.py bom.xlsx
    python scripts/readbom.py bom.xlsx --sheet "Sheet2" --json
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from typing import Dict, List, Optional

# Header names we recognise, mapped to a canonical field.
HEADER_HINTS = {
    "mpn": ("mpn", "manufacturer part", "mfr part", "mfg part", "manufacturer pn",
            "part number", "partnumber", "part no", "mfr. part"),
    "manufacturer": ("manufacturer", "mfr", "mfg", "brand", "make"),
    "lcsc": ("lcsc", "jlcpcb", "jlc part"),
    "distributor_pn": ("digikey", "digi-key", "mouser", "farnell", "rs part",
                       "supplier part", "distributor part"),
    "url": ("url", "link", "datasheet", "product page", "supplier link"),
    "value": ("value", "val"),
    "footprint": ("footprint", "package", "case", "package/case"),
    "designator": ("designator", "reference", "refdes", "ref"),
    "qty": ("qty", "quantity"),
    "description": ("description", "desc", "comment"),
}

_LCSC_RE = re.compile(r"\bC\d{3,}\b", re.IGNORECASE)


def _canonical(header: str) -> Optional[str]:
    h = str(header or "").strip().lower()
    if not h:
        return None
    for field, hints in HEADER_HINTS.items():
        for hint in hints:
            if h == hint or h.startswith(hint) or hint in h:
                return field
    return None


def _read_csv(path: str) -> List[List[str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [[(c or "").strip() for c in row] for row in csv.reader(fh, dialect)]


def _read_xlsx(path: str, sheet: Optional[str]) -> List[List[str]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is needed for .xlsx files:  python -m pip install openpyxl"
        ) from exc
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in row])
    wb.close()
    return rows


def _find_header(rows: List[List[str]]) -> int:
    """The header is the first row where at least two cells look like headers."""
    best, best_score = 0, 0
    for i, row in enumerate(rows[:30]):
        score = sum(1 for c in row if _canonical(c))
        if score > best_score:
            best, best_score = i, score
    return best if best_score >= 2 else 0


def read(path: str, sheet: Optional[str] = None) -> Dict[str, object]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        rows = _read_csv(path)
        sheets = []
    elif ext in (".xlsx", ".xlsm", ".xltx"):
        rows = _read_xlsx(path, sheet)
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    else:
        raise RuntimeError(f"unsupported file type {ext!r}; use .xlsx or .csv")

    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        return {"file": path, "sheets": sheets, "columns": {}, "parts": []}

    hdr_i = _find_header(rows)
    header = rows[hdr_i]
    mapping = {}
    for idx, cell in enumerate(header):
        field = _canonical(cell)
        if field and field not in mapping:
            mapping[field] = idx

    parts = []
    for row in rows[hdr_i + 1:]:
        entry: Dict[str, str] = {}
        for field, idx in mapping.items():
            if idx < len(row) and row[idx]:
                entry[field] = row[idx]
        if not entry:
            continue
        # An LCSC code often hides in a free-text cell rather than its own column.
        if "lcsc" not in entry:
            for cell in row:
                m = _LCSC_RE.search(cell or "")
                if m and "lcsc" in (cell or "").lower():
                    entry["lcsc"] = m.group(0).upper()
                    break
        if any(entry.get(k) for k in ("mpn", "lcsc", "distributor_pn", "url")):
            parts.append(entry)

    return {
        "file": path,
        "sheets": sheets,
        "header_row": hdr_i + 1,
        "columns": {k: header[v] for k, v in mapping.items()},
        "unmapped_columns": [c for i, c in enumerate(header)
                             if c and i not in mapping.values()],
        "parts": parts,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Read part identifiers from a BOM.")
    ap.add_argument("path")
    ap.add_argument("--sheet")
    ap.add_argument("--json", action="store_true")
    ap.add_argument("--limit", type=int, default=60)
    args = ap.parse_args()

    try:
        data = read(args.path, args.sheet)
    except (RuntimeError, OSError) as exc:
        print(f"ERROR: {exc}")
        return 1

    if args.json:
        print(json.dumps(data, indent=2))
        return 0

    print(f"  file          {data['file']}")
    if data["sheets"]:
        print(f"  sheets        {', '.join(data['sheets'])}")
    print(f"  header row    {data.get('header_row')}")
    print(f"  recognised    {data['columns']}")
    if data["unmapped_columns"]:
        print(f"  ignored       {data['unmapped_columns']}")
    parts = data["parts"]
    print(f"\n  {len(parts)} part rows:\n")
    for p in parts[: args.limit]:  # type: ignore[index]
        bits = [f"{k}={v}" for k, v in p.items()]
        print("    " + "  ".join(bits))
    if len(parts) > args.limit:  # type: ignore[arg-type]
        print(f"    ... and {len(parts) - args.limit} more")  # type: ignore[operator]
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
